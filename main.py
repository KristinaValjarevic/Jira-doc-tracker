import requests
from requests.auth import HTTPBasicAuth
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from config import username, password
from db_models import ReleaseNumber
from utils import get_session


doc_file = "C:\\Users\\kiki\\OneDrive\\Documentation sheet.xlsx"
issue_file = "C:\\Users\\kiki\\OneDrive\\Ticket sheet.xlsx"

doc_wb = openpyxl.load_workbook(doc_file)
issue_wb = openpyxl.load_workbook(issue_file)

session = get_session()
current_release_row = session.query(ReleaseNumber).first()
current_release = current_release_row.rn if current_release_row else None
session.close()


def add_row(sheet, row, start_at=1):
    cells_to_merge = []
    if len(sheet.merged_cells.ranges) > start_at:
        for merged_cells in sheet.merged_cells.ranges[start_at:]:
            cells_to_merge.append(move_cell_range(str(merged_cells)))
            sheet.unmerge_cells(str(merged_cells))
    sheet.insert_rows(row)
    if cells_to_merge:
        for cell_range in cells_to_merge:
            sheet.merge_cells(cell_range)


def insert_headers(release):
    for d_sheet in doc_wb:
        add_row(d_sheet, 2, start_at=0)
        add_row(d_sheet, 2, start_at=0)
        if d_sheet.title == "Special":
            d_sheet.merge_cells('A2:J2')
        else:
            d_sheet.merge_cells('A2:I2')
        cell = d_sheet['A2']
        cell.value = release
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="fff2cc")
        cell.alignment = Alignment(horizontal="center")
    doc_wb.save(doc_file)
    for i_sheet in issue_wb:
        add_row(i_sheet, 2, start_at=0)
        add_row(i_sheet, 2, start_at=0)
        i_sheet.merge_cells('A2:G2')
        cell = i_sheet['A2']
        cell.value = release
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="fff2cc")
        cell.alignment = Alignment(horizontal="center")
    issue_wb.save(issue_file)


def add_release_to_db(release):
    global current_release
    session = get_session()
    if current_release:
        release_row = session.query(ReleaseNumber).first()
        release_row.rn = release
    else:
        new_row = ReleaseNumber(rn=release)
        session.add(new_row)
    session.commit()
    session.close()
    current_release = release


def add_to_sheet(issue, team, row):
    sheet = doc_wb[team]
    # Checking whether to add the issue to the 'Special' sheet
    if issue['fields']['customfield_10029'] and "01" in issue['fields']['customfield_10029'] and \
            (not sheet[f'A{row}'].value or sheet[f'A{row}'].value.find("01") == -1):
        add_to_special(issue, team)
    # Document numbers and done documentation
    if issue['fields']['customfield_10029']:
        docs = ""
        done = ""
        for doc in issue['fields']['customfield_10029']:
            if docs:
                docs += f", {doc}"
                if issue['fields']['customfield_10031'] and doc in issue['fields']['customfield_10031']:
                    done += ", Yes"
                else:
                    done += ", No"
            else:
                docs = doc
                if issue['fields']['customfield_10031'] and doc in issue['fields']['customfield_10031']:
                    done = "Yes"
                else:
                    done = "No"
        sheet[f'A{row}'].value = docs
        sheet[f'B{row}'].value = done
    else:
        sheet[f'A{row}'].value = "-"
        sheet[f'B{row}'].value = "-"
    # Assignee
    if issue['fields']['assignee']:
        sheet[f'C{row}'].value = issue['fields']['assignee']['displayName']
    else:
        sheet[f'C{row}'].value = "-"
    # Jira ticket number
    sheet[f'D{row}'].value = issue['key']
    # Summary/description
    sheet[f'E{row}'].value = issue['fields']['summary']
    # Status of the ticket
    sheet[f'F{row}'].value = issue['fields']['status']['name']


def add_to_special(issue, team):
    sheet = doc_wb["Special"]
    # Find the first empty row
    row = 3
    while sheet[f'A{row}'].value:
        row += 1
    add_row(sheet, row)
    # Jira ticket number
    sheet[f'A{row}'].value = issue['key']
    # Summary/description
    sheet[f'B{row}'].value = issue['fields']['summary']
    # Status of the ticket
    sheet[f'C{row}'].value = issue['fields']['status']['name']
    # Done documentation
    if issue['fields']['customfield_10031'] and "01" in issue['fields']['customfield_10031']:
        sheet[f'D{row}'].value = "Yes"
    else:
        sheet[f'D{row}'].value = "No"
    # Team
    sheet[f'G{row}'] = team
    # Assignee
    if issue['fields']['assignee']:
        sheet[f'H{row}'].value = issue['fields']['assignee']['displayName']
    else:
        sheet[f'H{row}'].value = "-"


def update_sheet(team):
    # Collecting the list of issues already in the sheet
    sheet = doc_wb[team]
    columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
    row = 3
    sheet_issues = {}
    while sheet[f'D{row}'].value:
        if not sheet[f'D{row}'].font.strike:
            sheet_issues[sheet[f'D{row}'].value] = row
        row += 1
    # Iterating through the issues from the filter
    filter_issues = get_issues(f'{team} {current_release}', 'Yes')
    if filter_issues:
        for issue in filter_issues:
            if issue['key'] in sheet_issues:
                add_to_sheet(issue, team, sheet_issues[issue['key']])
                del sheet_issues[issue['key']]
            else:
                add_row(sheet, row)
                add_to_sheet(issue, team, row)
                row += 1
    # Highlighting the issues which were in the sheet but are not in the filter anymore
    if sheet_issues:
        for issue_id in sheet_issues.keys():
            for column in columns:
                cell = column + str(sheet_issues[issue_id])
                sheet[cell].fill = PatternFill("solid", fgColor="f4b084")
    doc_wb.save(doc_file)


def move_cell_range(cell_range):
    colon = cell_range.find(':')
    first = int(cell_range[1:colon]) + 1
    second = int(cell_range[colon + 2:]) + 1
    return cell_range[0] + str(first) + cell_range[colon:colon + 2] + str(second)


def api_request(url):
    auth = HTTPBasicAuth(username, password)
    headers = {
        "Accept": "application/json"
    }
    response = requests.request(
        "GET",
        url,
        headers=headers,
        auth=auth
    )
    return response.json()


def get_issues(filter_name, tagged, start=0):
    query = api_request(f'https://jira-doc-tracker.atlassian.net/rest/api/3/filter/search?filterName="{filter_name}"')
    jira_filter = query['values'][0]['self']
    jql = api_request(jira_filter)['jql']
    index = jql.find("ORDER BY") - 1
    new_jql = jql[0:index] + f' AND "Documentation[Checkboxes]" = {tagged} ' + jql[index:]
    issues = api_request(f"https://jira-doc-tracker.atlassian.net/rest/api/3/search?jql={new_jql}&fields"
                         f"=customfield_10029,customfield_10031,assignee,summary,status,customfield_10030,description"
                         f"&startAt={start}")
    if issues['total'] - issues['startAt'] <= issues['maxResults']:
        return issues['issues']
    else:
        start += issues['maxResults']
        return issues['issues'] + get_issues(filter_name, tagged, start=start)


if __name__ == "__main__":

    # Getting all issues in a filter

    result = get_issues(f"Team B {current_release}", "Yes")
    print(json.dumps(result, sort_keys=True, indent=4))


